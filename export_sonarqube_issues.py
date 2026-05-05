#!/usr/bin/env python3
"""
Export SonarQube Issues to Excel (per folder)
============================================
Fetches all issues from a SonarQube project via the REST API and exports them
to separate Excel files per folder (common, lmac, umac, unit_test).

Usage:
    python export_sonarqube_issues.py --token YOUR_SONAR_TOKEN
    python export_sonarqube_issues.py --token YOUR_SONAR_TOKEN --project ppe_riscv
    python export_sonarqube_issues.py --token YOUR_SONAR_TOKEN -o ./exports

Requires: requests, openpyxl
"""

from __future__ import annotations

import argparse
import os
import sys
from collections import defaultdict
from pathlib import Path

try:
    import requests
except ImportError:
    print("Error: 'requests' is required. Install with: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: 'openpyxl' is required. Install with: pip install openpyxl")
    sys.exit(1)

# Defaults from sonar-project.properties
DEFAULT_SONAR_URL = "https://sonarqube.silabs.net"
DEFAULT_PROJECT_KEY = "wifi_nwp_firmware"
DEFAULT_FOLDERS = ("common", "lmac", "umac", "unit_test")
PAGE_SIZE = 500


def _fetch_issues_page(
    api_url: str,
    auth: tuple[str, str],
    component_keys: str,
    page: int,
) -> tuple[list[dict], int]:
    """Fetch one page of issues for given component keys."""
    headers = {"Accept": "application/json"}
    params = {
        "componentKeys": component_keys,
        "ps": PAGE_SIZE,
        "p": page,
    }
    resp = requests.get(
        api_url,
        params=params,
        auth=auth,
        headers=headers,
        timeout=60,
    )
    resp.raise_for_status()
    data = resp.json()
    issues = data.get("issues", [])
    total = data.get("paging", {}).get("total", len(issues))
    return issues, total


def fetch_issues_for_component(
    sonar_url: str,
    project_key: str,
    token: str,
    component_pattern: str,
) -> list[dict]:
    """Fetch all issues for a component pattern (e.g. projectKey:wifi_nwp/lmac/%)."""
    base_url = sonar_url.rstrip("/")
    api_url = f"{base_url}/api/issues/search"
    auth = (token, "")

    all_issues = []
    page = 1

    while True:
        issues, total = _fetch_issues_page(api_url, auth, component_pattern, page)
        all_issues.extend(issues)
        if len(issues) == 0 or len(all_issues) >= total:
            break
        page += 1

    return all_issues


def get_folder_from_component(component: str, known_folders: tuple[str, ...]) -> str:
    """Extract folder from component path. Returns _other if no match."""
    path = component.split(":")[-1] if ":" in component else component
    path = path.replace("\\", "/")
    parts = [p for p in path.split("/") if p]
    for part in parts:
        if part in known_folders:
            return part
    return "_other"


def fetch_all_issues(sonar_url: str, project_key: str, token: str) -> list[dict]:
    """Fetch all issues from project (single API flow)."""
    base_url = sonar_url.rstrip("/")
    api_url = f"{base_url}/api/issues/search"
    auth = (token, "")
    all_issues = []
    page = 1
    while True:
        issues, total = _fetch_issues_page(api_url, auth, project_key, page)
        all_issues.extend(issues)
        if len(issues) == 0 or len(all_issues) >= total:
            break
        page += 1
    return all_issues


def fetch_all_issues_by_folder(
    sonar_url: str,
    project_key: str,
    token: str,
    folders: tuple[str, ...],
) -> dict[str, list[dict]]:
    """
    Fetch issues and group by folder. Tries component-key patterns first;
    if that returns nothing, falls back to fetch-all and group by path.
    """
    base_url = sonar_url.rstrip("/")
    api_url = f"{base_url}/api/issues/search"
    auth = (token, "")

    grouped: dict[str, list[dict]] = {}
    seen_keys: set[str] = set()

    # Try fetching per folder via component pattern
    patterns_to_try = (
        (f"{project_key}:wifi_nwp/{{folder}}/%", "wifi_nwp/<folder>/%"),
        (f"{project_key}:{{folder}}/%", "<folder>/%"),
    )
    for folder in folders:
        folder_issues = []
        for pattern_tpl, pattern_desc in patterns_to_try:
            component_pattern = pattern_tpl.format(folder=folder)
            print(f"  Fetching {folder} (pattern: {pattern_desc})...")
            all_issues = []
            page = 1
            while True:
                issues, total = _fetch_issues_page(api_url, auth, component_pattern, page)
                for issue in issues:
                    key = issue.get("key", "")
                    if key and key not in seen_keys:
                        seen_keys.add(key)
                        all_issues.append(issue)
                if len(issues) == 0 or len(all_issues) >= total:
                    break
                page += 1
            if all_issues:
                folder_issues = all_issues
                print(f"    -> {len(all_issues)} issues")
                break
        grouped[folder] = folder_issues

    # If no folder-specific fetches worked, fetch all and group by path
    total_fetched = sum(len(v) for v in grouped.values())
    if total_fetched == 0:
        print("  Component patterns returned no issues. Fetching all and grouping by path...")
        all_issues = fetch_all_issues(sonar_url, project_key, token)
        grouped = defaultdict(list)
        for issue in all_issues:
            folder = get_folder_from_component(issue.get("component", ""), folders)
            grouped[folder].append(issue)
        grouped = dict(grouped)

        # Show sample component paths when using fallback (helps debug path format)
        if all_issues:
            samples = list({i.get("component", "") for i in all_issues[:15]})
            print("  Sample component values (path format):")
            for s in samples[:8]:
                print(f"    {s}")
    else:
        # Fetch remaining (other) - issues not in any folder
        print("  Fetching remaining issues (other)...")
        other_issues = []
        page = 1
        while True:
            issues, total = _fetch_issues_page(api_url, auth, project_key, page)
            for issue in issues:
                key = issue.get("key", "")
                if key and key not in seen_keys:
                    seen_keys.add(key)
                    other_issues.append(issue)
            if len(issues) == 0 or len(other_issues) >= total:
                break
            page += 1
        if other_issues:
            grouped["_other"] = other_issues
            print(f"    -> {len(other_issues)} issues")

    return grouped


def issue_to_row(issue: dict) -> list:
    """Convert a SonarQube issue dict to a flat row for Excel."""
    component = issue.get("component", "")
    # component can be like "wifi_nwp_firmware:wifi_nwp/lmac/..."
    component_short = component.split(":")[-1] if ":" in component else component

    return [
        issue.get("key", ""),
        issue.get("rule", ""),
        issue.get("severity", ""),
        issue.get("type", ""),
        issue.get("message", ""),
        component_short,
        issue.get("line"),
        issue.get("status", ""),
        issue.get("resolution", "") or "-",
        issue.get("effort", "") or "-",
        issue.get("debt", "") or "-",
        issue.get("creationDate", ""),
        issue.get("updateDate", ""),
        issue.get("hash", "") or "-",
    ]


def export_to_excel(issues: list[dict], output_path: Path) -> None:
    """Write issues to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SonarQube Issues"

    headers = [
        "Key",
        "Rule",
        "Severity",
        "Type",
        "Message",
        "Component",
        "Line",
        "Status",
        "Resolution",
        "Effort",
        "Debt",
        "Creation Date",
        "Update Date",
        "Hash",
    ]

    # Header row with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, issue in enumerate(issues, start=2):
        row_data = issue_to_row(issue)
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export SonarQube issues to an Excel file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--url",
        type=str,
        default=os.environ.get("SONAR_HOST_URL", DEFAULT_SONAR_URL),
        help=f"SonarQube server URL (default: {DEFAULT_SONAR_URL})",
    )
    parser.add_argument(
        "--project",
        type=str,
        default=os.environ.get("SONAR_PROJECT_KEY", DEFAULT_PROJECT_KEY),
        help=f"Project key (default: {DEFAULT_PROJECT_KEY})",
    )
    parser.add_argument(
        "-t",
        "--token",
        type=str,
        default=os.environ.get("SONAR_TOKEN"),
        help="SonarQube token (or set SONAR_TOKEN env var)",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=str,
        default=".",
        help="Output directory for Excel files (default: current directory)",
    )
    parser.add_argument(
        "--folders",
        type=str,
        default=",".join(DEFAULT_FOLDERS),
        help=f"Comma-separated folder names to split by (default: {','.join(DEFAULT_FOLDERS)})",
    )

    args = parser.parse_args()

    if not args.token:
        parser.error("SonarQube token is required. Use --token or set SONAR_TOKEN.")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    folders = tuple(f.strip() for f in args.folders.split(",") if f.strip())

    print(f"Fetching issues from {args.url} for project '{args.project}'...")
    grouped = fetch_all_issues_by_folder(args.url, args.project, args.token, folders)

    total_issues = sum(len(issues) for issues in grouped.values())
    if total_issues == 0:
        print("No issues found.")
        return

    safe_project = args.project.replace(" ", "_").replace("/", "_")

    for folder_name, folder_issues in sorted(grouped.items()):
        if not folder_issues:
            continue
        out_file = output_dir / f"sonarqube_issues_{safe_project}_{folder_name}.xlsx"
        export_to_excel(folder_issues, out_file)
        print(f"  {folder_name}: {len(folder_issues)} issues -> {out_file.name}")


if __name__ == "__main__":
    main()
